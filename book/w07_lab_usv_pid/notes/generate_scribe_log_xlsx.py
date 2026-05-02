#!/usr/bin/env python3
"""Generate scribe_log.xlsx — Lab 2 USV PID field data-collection template.

Setup (one-time, in the project venv):
  utils/ardu_utils/.venv/bin/pip install openpyxl

Run:
  utils/ardu_utils/.venv/bin/python book/w07_lab_usv_pid/notes/generate_scribe_log_xlsx.py

Output: book/w07_lab_usv_pid/notes/scribe_log.xlsx
The xlsx is committed in the repo, so you only need to re-run this if
you edit the layout/columns. For day-to-day edits, just open the xlsx
in Excel or LibreOffice Calc.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "scribe_log.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
TITLE_FONT     = Font(name="Calibri", size=16, bold=True)
SECTION_FONT   = Font(name="Calibri", size=12, bold=True)
HEADER_FONT    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
GROUP_FONT     = Font(name="Calibri", size=11, bold=True)
LABEL_FONT     = Font(name="Calibri", size=10, italic=True)
PARAM_FONT     = Font(name="Consolas", size=10)

HEADER_FILL    = PatternFill("solid", fgColor="305496")   # dark blue
GROUP_FILL_S   = PatternFill("solid", fgColor="DDEBF7")   # surge group (light blue)
GROUP_FILL_Y   = PatternFill("solid", fgColor="E2EFDA")   # yaw group (light green)
ALT_FILL       = PatternFill("solid", fgColor="F2F2F2")   # alternating row
LEGEND_FILL    = PatternFill("solid", fgColor="FFF2CC")   # legend tab

THIN  = Side(border_style="thin", color="999999")
THICK = Side(border_style="medium", color="305496")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER_THIN


def style_data(cell, fill=None):
    cell.alignment = CENTER
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill


# ── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()

# ───────────────────────────── Sheet 1: Run Log ─────────────────────────────
ws = wb.active
ws.title = "Run Log"

# Title
ws.merge_cells("A1:Q1")
c = ws["A1"]
c.value = "Lab 2 Scribe Log — USV PID Control"
c.font = TITLE_FONT
c.alignment = CENTER

# Header info block
header_rows = [
    ("Date:",    "B3:D3", "Vehicle:", "F3:H3"),
    ("Team:",    "B4:D4", "",         None),
    ("Pilot:",   "B5:D5", "Planner:", "F5:H5"),
    ("Scribe:",  "B6:D6", "",         None),
]
for label, val_range, label2, val2_range in header_rows:
    row = int(val_range[1])
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = SECTION_FONT
    ws.merge_cells(val_range)
    ws[val_range.split(":")[0]].border = BORDER_THIN
    if label2:
        ws[f"E{row}"] = label2
        ws[f"E{row}"].font = SECTION_FONT
        ws.merge_cells(val2_range)
        ws[val2_range.split(":")[0]].border = BORDER_THIN

# Pre-run checklist
ws["A8"] = "Pre-Run Checklist (verify before every arming)"
ws["A8"].font = SECTION_FONT
checklist = [
    "Mode set to ACRO before arming",
    "Parameter file saved in Mission Planner before arming (timestamp in filename)",
    "Scribe records time and log number AT arming",
    "Same throttle/rudder step fraction across Run 1 and Run 2",
    "Surge and yaw trials in separate log files (disarm between channels)",
]
for i, item in enumerate(checklist, start=9):
    ws[f"A{i}"] = "☐"
    ws[f"A{i}"].alignment = CENTER
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=17)
    ws.cell(row=i, column=2).value = item
    ws.cell(row=i, column=2).alignment = LEFT

# Run log table
table_start = 16
ws[f"A{table_start-1}"] = "Run Log — one row per arming. Dash (—) means N/A for this trial type."
ws[f"A{table_start-1}"].font = SECTION_FONT

# Two-row header: group header + column header
group_row = table_start
col_row   = table_start + 1

# Six leftmost columns are merged across two rows; group cells span 5 cols each
left_headers = ["Run", "Time (PDT)", "ARM", "ACRO", "Log #", "Trial"]
for i, h in enumerate(left_headers, start=1):
    ws.merge_cells(start_row=group_row, start_column=i, end_row=col_row, end_column=i)
    cell = ws.cell(row=group_row, column=i, value=h)
    style_header(cell)

# Group header for surge (cols 7-11) and yaw (cols 12-16)
ws.merge_cells(start_row=group_row, start_column=7, end_row=group_row, end_column=11)
gs = ws.cell(row=group_row, column=7, value="Surge PID")
gs.font = GROUP_FONT
gs.fill = GROUP_FILL_S
gs.alignment = CENTER
gs.border = BORDER_THIN

ws.merge_cells(start_row=group_row, start_column=12, end_row=group_row, end_column=16)
gy = ws.cell(row=group_row, column=12, value="Yaw-rate PID")
gy.font = GROUP_FONT
gy.fill = GROUP_FILL_Y
gy.alignment = CENTER
gy.border = BORDER_THIN

# Notes column header (col 17, spans both rows)
ws.merge_cells(start_row=group_row, start_column=17, end_row=col_row, end_column=17)
nc = ws.cell(row=group_row, column=17, value="Notes")
style_header(nc)

# Sub-headers under group headers
sub_headers_surge = ["P", "I", "D", "FF", "Max"]
sub_headers_yaw   = ["P", "I", "D", "FF", "Max"]
for i, h in enumerate(sub_headers_surge, start=7):
    cell = ws.cell(row=col_row, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER_THIN
for i, h in enumerate(sub_headers_yaw, start=12):
    cell = ws.cell(row=col_row, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER_THIN

# Data rows: alternating Surge / Yaw trials, 12 rows
data_start = col_row + 1
trial_pattern = ["Surge", "Yaw"] * 5 + ["", ""]   # 12 rows
for r, trial in enumerate(trial_pattern):
    row = data_start + r
    fill = ALT_FILL if r % 2 == 0 else None
    # Run number
    c = ws.cell(row=row, column=1, value=r + 1)
    style_data(c, fill)
    # Time, ARM, ACRO, Log #, Trial
    for col in (2, 5):
        style_data(ws.cell(row=row, column=col), fill)
    for col in (3, 4):  # ARM, ACRO
        c = ws.cell(row=row, column=col, value="☐")
        style_data(c, fill)
    c = ws.cell(row=row, column=6, value=trial)
    style_data(c, fill)
    # Surge cols 7-11 — empty for Yaw, fillable for Surge
    for col in range(7, 12):
        c = ws.cell(row=row, column=col, value="—" if trial == "Yaw" else None)
        style_data(c, fill)
    # Yaw cols 12-16 — empty for Surge
    for col in range(12, 17):
        c = ws.cell(row=row, column=col, value="—" if trial == "Surge" else None)
        style_data(c, fill)
    # Notes
    c = ws.cell(row=row, column=17, value=None)
    c.alignment = LEFT
    c.border = BORDER_THIN
    if fill:
        c.fill = fill

# Set column widths
widths = {
    "A": 5, "B": 12, "C": 6, "D": 6, "E": 8, "F": 8,
    "G": 7, "H": 7, "I": 7, "J": 7, "K": 7,
    "L": 7, "M": 7, "N": 7, "O": 7, "P": 7,
    "Q": 32,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Set row heights for the data rows so there's room to write
for r in range(data_start, data_start + len(trial_pattern)):
    ws.row_dimensions[r].height = 28

# Param file record block
pf_start = data_start + len(trial_pattern) + 2
ws.cell(row=pf_start, column=1, value="Parameter File Record").font = SECTION_FONT
ws.merge_cells(start_row=pf_start, start_column=1, end_row=pf_start, end_column=17)

pf_header_row = pf_start + 1
ws.cell(row=pf_header_row, column=1, value="#")
ws.merge_cells(start_row=pf_header_row, start_column=2, end_row=pf_header_row, end_column=6)
ws.cell(row=pf_header_row, column=2, value="Filename")
ws.merge_cells(start_row=pf_header_row, start_column=7, end_row=pf_header_row, end_column=17)
ws.cell(row=pf_header_row, column=7, value="Changes from previous")
for col in (1, 2, 7):
    style_header(ws.cell(row=pf_header_row, column=col))

for i in range(6):
    row = pf_header_row + 1 + i
    fill = ALT_FILL if i % 2 == 0 else None
    c = ws.cell(row=row, column=1, value=i + 1)
    style_data(c, fill)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=2, value=("Baseline" if i == 0 else None))
    cell.alignment = LEFT
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=17)
    cell = ws.cell(row=row, column=7)
    cell.alignment = LEFT
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill
    ws.row_dimensions[row].height = 24

# Freeze panes — keep title + run-log header visible while scrolling
ws.freeze_panes = ws.cell(row=col_row + 1, column=1)

# ───────────────────────────── Sheet 2: Legend ──────────────────────────────
ws2 = wb.create_sheet("Legend")
ws2["A1"] = "Parameter Legend"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:C1")
ws2["A1"].alignment = CENTER

ws2["A3"] = "View / set in Mission Planner: Config → Full Parameter List"
ws2["A3"].font = LABEL_FONT
ws2.merge_cells("A3:C3")

legend = [
    ("SP_P",   "ATC_SPEED_P",       "Surge speed proportional gain"),
    ("SP_I",   "ATC_SPEED_I",       "Surge speed integral gain"),
    ("SP_D",   "ATC_SPEED_D",       "Surge speed derivative gain"),
    ("SP_FF",  "ATC_SPEED_FF",      "Surge speed feed-forward"),
    ("A_MAX",  "ATC_ACCEL_MAX",     "Speed cmd rate limit (m/s²); 0 = off"),
    ("YR_P",   "ATC_STR_RAT_P",     "Yaw-rate proportional gain"),
    ("YR_I",   "ATC_STR_RAT_I",     "Yaw-rate integral gain"),
    ("YR_D",   "ATC_STR_RAT_D",     "Yaw-rate derivative gain"),
    ("YR_FF",  "ATC_STR_RAT_FF",    "Yaw-rate feed-forward"),
    ("YA_MAX", "ATC_STR_ACC_MAX",   "Yaw-rate cmd rate limit (deg/s²); 0 = off"),
]
ws2.append(["Column", "ArduPilot Parameter", "Description"])
for col in range(1, 4):
    style_header(ws2.cell(row=ws2.max_row, column=col))
for short, full, desc in legend:
    ws2.append([short, full, desc])
    r = ws2.max_row
    fill = ALT_FILL if (r % 2 == 0) else None
    for col, font in zip(range(1, 4), (PARAM_FONT, PARAM_FONT, None)):
        cell = ws2.cell(row=r, column=col)
        cell.border = BORDER_THIN
        cell.alignment = LEFT
        if font:
            cell.font = font
        if fill:
            cell.fill = fill

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 50

# Freeze legend header
ws2.freeze_panes = "A5"

# ── Save ────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Wrote {OUT}")
